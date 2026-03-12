#!/usr/bin/env python3
"""
Extract embedded images from an Excel sheet and rename them by id_codigo.

Expected layout (customizable by args):
- Photo column: H
- Code column: I

For each embedded image anchored in the photo column row, this script reads the
code from the code column in the same row and writes a file named <id_codigo>.<ext>.
"""

from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import load_workbook


def sanitize_code(value: object) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    # Keep only safe chars for filenames.
    text = re.sub(r"[^A-Za-z0-9_-]", "", text)
    if not text:
        return None
    return text.upper()


def image_extension(image_obj) -> str:
    # openpyxl images often have a path like '/xl/media/image1.jpeg'
    path = getattr(image_obj, "path", "") or ""
    suffix = Path(path).suffix.lower().lstrip(".")
    if suffix in {"jpg", "jpeg", "png", "gif", "bmp", "webp"}:
        return "jpg" if suffix == "jpeg" else suffix
    return "png"


def anchored_row(image_obj) -> Optional[int]:
    anchor = getattr(image_obj, "anchor", None)
    if anchor is None:
        return None
    marker = getattr(anchor, "_from", None)
    if marker is None:
        return None
    # openpyxl row index is 0-based.
    return int(marker.row) + 1


def anchored_col(image_obj) -> Optional[int]:
    anchor = getattr(image_obj, "anchor", None)
    if anchor is None:
        return None
    marker = getattr(anchor, "_from", None)
    if marker is None:
        return None
    # openpyxl col index is 0-based.
    return int(marker.col) + 1


def ensure_unique_name(base_name: str, ext: str, used: set[str]) -> str:
    candidate = f"{base_name}.{ext}"
    if candidate not in used:
        used.add(candidate)
        return candidate

    idx = 2
    while True:
        candidate = f"{base_name}_{idx}.{ext}"
        if candidate not in used:
            used.add(candidate)
            return candidate
        idx += 1


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract and rename Excel embedded photos by id_codigo."
    )
    parser.add_argument("--xlsx", required=True, help="Path to .xlsx file")
    parser.add_argument("--sheet", default=None, help="Sheet name (default: active)")
    parser.add_argument(
        "--photo-col",
        default="H",
        help="Photo column letter where images are anchored (default: H)",
    )
    parser.add_argument(
        "--code-col",
        default="I",
        help="Code column letter with id_codigo (default: I)",
    )
    parser.add_argument(
        "--out-dir",
        default="personal_fotos_output",
        help="Output folder for renamed files",
    )
    parser.add_argument(
        "--manifest",
        default="manifest_personal_fotos.csv",
        help="CSV manifest filename",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()

    xlsx_path = Path(args.xlsx).expanduser().resolve()
    out_dir = Path(args.out_dir).expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    manifest_path = out_dir / args.manifest

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb.active

    photo_col_idx = ws[args.photo_col + "1"].column
    code_col = args.code_col.upper()

    images: Iterable = getattr(ws, "_images", [])

    used_names: set[str] = set()
    extracted = 0
    skipped = 0

    with manifest_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["row", "id_codigo", "file_name", "status", "detail"])

        for img in images:
            row = anchored_row(img)
            col = anchored_col(img)

            if row is None or col is None:
                skipped += 1
                writer.writerow(["", "", "", "skipped", "image without anchor"])
                continue

            # Ignore images anchored outside the expected photo column.
            if col != photo_col_idx:
                skipped += 1
                writer.writerow([row, "", "", "skipped", f"image not in photo column {args.photo_col}"])
                continue

            code_raw = ws[f"{code_col}{row}"].value
            code = sanitize_code(code_raw)
            if not code:
                skipped += 1
                writer.writerow([row, str(code_raw or ""), "", "skipped", "missing/invalid id_codigo"])
                continue

            ext = image_extension(img)
            file_name = ensure_unique_name(code, ext, used_names)
            output_path = out_dir / file_name

            try:
                output_path.write_bytes(img._data())
            except Exception as ex:  # noqa: BLE001
                skipped += 1
                writer.writerow([row, code, file_name, "error", str(ex)])
                continue

            extracted += 1
            writer.writerow([row, code, file_name, "ok", ""])

    print(f"Workbook: {xlsx_path}")
    print(f"Sheet: {ws.title}")
    print(f"Output folder: {out_dir}")
    print(f"Extracted: {extracted}")
    print(f"Skipped/errors: {skipped}")
    print(f"Manifest: {manifest_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
